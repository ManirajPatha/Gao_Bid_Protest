import argparse
import json
import os
import random
import re
import time
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter, Retry

try:
    import cloudscraper
except Exception:
    cloudscraper = None

SEARCH_URL_DEFAULT = (
    "https://www.gao.gov/search?keyword=Bid%20Protest%20Decisions"
    "&facets_query=&f%5B0%5D=ctype_search%3ABid%20Protest"
    "&f%5B1%5D=ctype_search%3ABid%20Protest%20Decision"
)

def human_sleep(a: float = 0.35, b: float = 0.9) -> None:
    time.sleep(a + (b - a) * random.random())

def sanitize_for_excel(s: Optional[str]) -> str:
    if s is None:
        return ""
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)

def sanitize_for_json(s: Optional[str]) -> str:
    if s is None:
        return ""
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)

def normalize_text(txt: str) -> str:
    if not txt:
        return ""
    txt = txt.replace("\r", "")
    txt = re.sub(r"(\w)-\n(\w)", r"\1\2", txt)
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    return txt.strip()

def clean_report_text(txt: str) -> str:
    if not txt:
        return ""
    lines = [l.rstrip() for l in txt.split("\n")]
    drop_patterns = [
        r"^441 G St\. N\.W\.$",
        r"^Washington, DC\b.*$",
        r"^Comptroller General\b.*$",
        r"^of the United States$",
        r"^U\.?S\.? Government Accountability Office$",
        r"^www\.gao\.gov$",
        r"^Page\s+\d+\s*$",
        r"^B-\d{4,7}(\.\d+)?\s*$",
        r"^\s*~+\s*$|^\s*–+\s*$|^\s*-{2,}\s*$",
    ]
    drop_rx = re.compile("|".join(drop_patterns), re.IGNORECASE)
    kept = [l for l in lines if not drop_rx.search(l)]
    txt2 = "\n".join(kept)
    txt2 = re.sub(r"\n\s*B-\d{4,7}(\.\d+)?\s*\n", "\n", txt2)
    return normalize_text(txt2)

KNOWN_ORDER = [
    "DIGEST",
    "BACKGROUND",
    "DISCUSSION",
    "DECISION",
    "CONCLUSION",
    "RECOMMENDATION",
    "CONCLUSIONS",
    "RECOMMENDATIONS",
]
CAPS_LINE = r"(?m)^(?P<cap>[A-Z0-9][A-Z0-9\s’'()/\-,.:;]{3,40})$"

def split_sections(full_text: str) -> Dict[str, str]:
    ft = normalize_text(full_text)
    if not ft:
        return {}
    found_positions: Dict[str, int] = {}
    for h in KNOWN_ORDER:
        m = re.search(fr"(?m)^{re.escape(h)}\s*$", ft)
        if m:
            found_positions[h] = m.start()
    for m in re.finditer(CAPS_LINE, ft):
        cap = m.group("cap").strip()
        if cap.upper() == cap and cap not in ("U N I T E D  S T A T E S",):
            found_positions.setdefault(cap, m.start())
    if not found_positions:
        return {"Full Report Text": ft}
    ordered = [h for h, _ in sorted(found_positions.items(), key=lambda kv: kv[1])]
    rx = re.compile("(?m)^(" + "|".join(re.escape(h) for h in ordered) + r")\s*$")
    marks = [(m.group(0).strip(), m.start(), m.end()) for m in rx.finditer(ft)]
    out: Dict[str, str] = {}
    for i, (name, st, en) in enumerate(marks):
        body = ft[en:(marks[i + 1][1] if i + 1 < len(marks) else len(ft))].strip()
        out[name] = body
    return out

_UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.85 Safari/537.36",
]

def _rand_ua() -> str:
    return random.choice(_UAS)

def build_session() -> requests.Session:
    if cloudscraper is not None:
        s = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
    else:
        s = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=0.7,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=frozenset(["GET"]),
            raise_on_status=False,
        )
        s.mount("https://", HTTPAdapter(max_retries=retries))
        s.mount("http://", HTTPAdapter(max_retries=retries))
    ua = _rand_ua()
    s.headers.update({
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
    })
    for warm in ["https://www.gao.gov/", "https://www.gao.gov/search"]:
        try:
            s.get(warm, timeout=30)
            time.sleep(0.6)
        except Exception:
            pass
    return s

def get_html(session: requests.Session, url: str) -> Optional[BeautifulSoup]:
    def _fetch(u: str) -> requests.Response:
        hdrs = {}
        if "/search" in u:
            hdrs["Referer"] = "https://www.gao.gov/"
        elif "/products/" in u:
            hdrs["Referer"] = "https://www.gao.gov/search"
        return session.get(u, headers=hdrs, timeout=60)
    try:
        r = _fetch(url)
        if r.status_code == 200:
            return BeautifulSoup(r.text, "lxml")
        if r.status_code == 403:
            session.headers["User-Agent"] = _rand_ua()
            time.sleep(1.0)
            r2 = _fetch(url)
            if r2.status_code == 200:
                return BeautifulSoup(r2.text, "lxml")
            new_session = build_session()
            time.sleep(1.0)
            r3 = new_session.get(url, timeout=60)
            if r3.status_code == 200:
                session.cookies = new_session.cookies
                session.headers.update(new_session.headers)
                return BeautifulSoup(r3.text, "lxml")
            print(f"[WARN] GET {url} -> {r3.status_code} after rebuild")
            return None
        print(f"[WARN] GET {url} -> {r.status_code}")
        return None
    except Exception as e:
        print(f"[WARN] request failed {url}: {e}")
        return None

def collect_result_links_from_page(soup: BeautifulSoup) -> List[str]:
    links: List[str] = []
    seen: Set[str] = set()
    main = soup.select_one("main") or soup
    for a in main.select("a[href*='/products/']"):
        href = a.get("href", "")
        title = (a.get_text(strip=True) or "")
        if href and "/products/" in href and title and href not in seen:
            if href.startswith("/"):
                href = "https://www.gao.gov" + href
            links.append(href)
            seen.add(href)
    return links

def get_next_page(soup: BeautifulSoup) -> Optional[str]:
    a = soup.select_one("a[rel='next']") or soup.find("a", string=lambda t: t and "Next" in t)
    if not a:
        return None
    href = a.get("href", "")
    if not href:
        return None
    if href.startswith("/"):
        href = "https://www.gao.gov" + href
    return href

def get_title_file_date_from_doc(soup: BeautifulSoup) -> Tuple[str, str, str]:
    title = ""
    file_no = ""
    date = ""
    h1 = soup.find("h1")
    if h1:
        title = h1.get_text(strip=True)
    main_text = (soup.select_one("main") or soup).get_text(" ", strip=True)
    m = re.search(r"B-\d{4,7}(?:\.\d+)?", main_text)
    if m:
        file_no = m.group(0)
    m = re.search(
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}",
        main_text,
    )
    if m:
        date = m.group(0)
    return title, file_no, date

def extract_expanded_decision_text(soup: BeautifulSoup) -> str:
    container = soup.select_one("div.field__item[data-readmore]") or soup.select_one("div[data-readmore]")
    if container:
        for br in container.find_all(["br"]):
            br.replace_with("\n")
        return container.get_text("\n", strip=True)
    for block in soup.select("div.field__item"):
        if block.find("p") and "Decision" in block.get_text(" ", strip=True)[:50]:
            for br in block.find_all(["br"]):
                br.replace_with("\n")
            return block.get_text("\n", strip=True)
    return (soup.select_one("main") or soup).get_text("\n", strip=True)

def scrape_item_with_bs(session: requests.Session, url: str) -> dict:
    soup = get_html(session, url)
    if soup is None:
        return {
            "base": {"url": url, "title": "", "file_number": "", "date": ""},
            "pdf_pages": None,
            "full_text": "",
            "sections": {},
        }
    title, file_no, date = get_title_file_date_from_doc(soup)
    raw = extract_expanded_decision_text(soup)
    cleaned = clean_report_text(raw)
    sections = split_sections(cleaned)
    return {
        "base": {"url": url, "title": title, "file_number": file_no, "date": date},
        "pdf_pages": None,
        "full_text": normalize_text(cleaned),
        "sections": sections,
    }

def write_outputs(records: List[dict], out_csv: str, out_xlsx: str) -> None:
    upload_rows = []
    for r in records:
        base = r["base"]
        sections_clean = {
            k: sanitize_for_json(v)
            for k, v in r["sections"].items()
            if v and v.strip() and k != "Full Report Text"
        }
        meta = {
            "file_number": base.get("file_number") or "",
            "title": base.get("title") or "",
            "date": base.get("date") or "",
            "pdf_pages": r.get("pdf_pages"),
            "url": base.get("url") or "",
            "sections": sections_clean,
        }
        meta_json = json.dumps(meta, ensure_ascii=False, separators=(",", ":"))
        full_txt = sanitize_for_json(r.get("full_text", ""))
        upload_rows.append(
            {"protest_id": "", "file_metadata": meta_json, "file_content": full_txt}
        )
    upload_df = pd.DataFrame(upload_rows, columns=["protest_id", "file_metadata", "file_content"])
    upload_df.to_csv(out_csv, index=False, encoding="utf-8")
    print(f"[OK] DB-ready file written → {out_csv}  (rows: {len(upload_rows)})")
    if out_xlsx:
        KNOWN_ORDER_PREF = [
            "DIGEST", "BACKGROUND", "DISCUSSION", "DECISION", "CONCLUSION", "RECOMMENDATION"
        ]
        meta_cols = ["file_number", "title", "date", "pdf_pages", "url"]
        rows = []
        sheet_map = []
        for idx, r in enumerate(records, 1):
            base = r["base"]
            row = {
                "file_number": sanitize_for_excel(base.get("file_number", "")),
                "title": sanitize_for_excel(base.get("title", "")),
                "date": sanitize_for_excel(base.get("date", "")),
                "pdf_pages": r.get("pdf_pages", "") or "",
                "url": base.get("url", ""),
            }
            secs = r.get("sections", {})
            for k in KNOWN_ORDER_PREF:
                row[k] = sanitize_for_excel(secs.get(k, ""))
            sheet_name = re.sub(
                r"[\\/*?:\[\]]",
                "_",
                row["file_number"] or row["title"] or f"Item {idx}",
            )[:31] or f"Item {idx}"
            col_title = (
                f"GAO Bid Protest Decision – "
                f"{row['title'] or row['file_number'] or sheet_name} – Complete Text"
            )
            sheet_map.append((sheet_name, col_title, sanitize_for_excel(r.get("full_text", ""))))
            rows.append(row)
        master_cols = meta_cols + KNOWN_ORDER_PREF
        df = pd.DataFrame(rows, columns=master_cols)
        from openpyxl.styles import Alignment
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xlw:
            df.to_excel(xlw, sheet_name="Master", index=False)
            ws = xlw.book["Master"]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            width_map = {"file_number": 14, "title": 45, "date": 14, "pdf_pages": 10, "url": 36}
            for col in ws.iter_cols(min_row=1, max_row=1):
                header = col[0].value
                ws.column_dimensions[col[0].column_letter].width = width_map.get(
                    header, 50 if header in KNOWN_ORDER_PREF else 22
                )
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if ws.cell(1, cell.column).value in KNOWN_ORDER_PREF:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top")
            for sheet_name, col_title, text in sheet_map:
                pd.DataFrame({col_title: [text]}).to_excel(xlw, sheet_name=sheet_name, index=False)
                wsi = xlw.book[sheet_name]
                wsi.freeze_panes = "A2"
                wsi.column_dimensions["A"].width = 120
                if wsi.max_row >= 2:
                    wsi["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        print(f"[OK] Review workbook written → {out_xlsx}  (items: {len(records)})")

def run(search_url: str, out_csv: str, out_xlsx: str, max_pages: int, upto: int):
    session = build_session()
    records: List[dict] = []
    processed = 0
    page_url = search_url
    page_num = 1
    try:
        while page_url:
            soup = get_html(session, page_url)
            if soup is None:
                print(f"[WARN] Skipping page (no HTML): {page_url}")
                break
            links = collect_result_links_from_page(soup)
            if not links:
                print("[INFO] No product links found on page.")
                break
            for url in links:
                if upto and processed >= upto:
                    print(f"[INFO] Reached --upto limit ({upto}). Stopping.")
                    raise StopIteration
                rec = scrape_item_with_bs(session, url)
                records.append(rec)
                processed += 1
                try:
                    write_outputs(records, out_csv, out_xlsx)
                except PermissionError as pe:
                    print(f"[WARN] Excel locked, writing fallback: {pe}")
                    base, ext = os.path.splitext(out_xlsx)
                    write_outputs(records, out_csv, f"{base}_partial{ext}")
                human_sleep()
            page_num += 1
            if max_pages and page_num > max_pages:
                break
            nxt = get_next_page(soup)
            if not nxt:
                break
            page_url = nxt
            human_sleep(0.6, 1.2)
    except StopIteration:
        pass
    except KeyboardInterrupt:
        print("[WARN] Interrupted by user. Saving partial results.")
    except Exception as e:
        print(f"[WARN] Unexpected error: {e}")
    finally:
        try:
            write_outputs(records, out_csv, out_xlsx)
        except PermissionError as pe:
            print(f"[WARN] Final Excel write locked; writing fallback: {pe}")
            base, ext = os.path.splitext(out_xlsx)
            write_outputs(records, out_csv, f"{base}_partial{ext}")

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="GAO Bid Protests (BeautifulSoup) → DB-ready CSV/XLSX from View Decision HTML."
    )
    ap.add_argument("--url", default=SEARCH_URL_DEFAULT)
    ap.add_argument("--out-csv", default="goa_protest_file_upload.csv")
    ap.add_argument("--out-xlsx", default="gao_bid_protests.xlsx")
    ap.add_argument("--max-pages", type=int, default=0)
    ap.add_argument("--upto", type=int, default=0, help="Stop after scraping N rows (0 = no limit)")
    args = ap.parse_args()
    run(
        search_url=args.url,
        out_csv=args.out_csv,
        out_xlsx=args.out_xlsx,
        max_pages=args.max_pages,
        upto=args.upto,
    )
